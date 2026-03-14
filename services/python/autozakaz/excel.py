from __future__ import annotations

import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import SaleRow, Settings


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def parse_period(raw: Any) -> tuple[str | None, str | None, int]:
    text = str(raw or "")
    match = re.search(r"с\s+(\d{2}\.\d{2}\.\d{4})\s+по\s+(\d{2}\.\d{2}\.\d{4})", text)
    if not match:
        return None, None, 1
    start = datetime.strptime(match.group(1), "%d.%m.%Y").date()
    end = datetime.strptime(match.group(2), "%d.%m.%Y").date()
    return start.isoformat(), end.isoformat(), max((end - start).days + 1, 1)


def normalize_barcode(value: Any) -> str:
    if value in (None, ""):
        return ""

    if isinstance(value, bool):
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if math.isnan(value):
            return ""
        return str(int(value))

    text = str(value).strip()

    # если это строка вида "4610000000380.0"
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]

    return re.sub(r"\D", "", text)


def read_iiko_sales(path: str | Path) -> tuple[list[SaleRow], dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    ws = workbook.active

    header_row = None
    barcode_col = None
    name_col = None
    qty_col = None

    for row_idx in range(1, min(ws.max_row, 40) + 1):
        values = [normalize_header(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        current_barcode = None
        current_name = None
        current_qty = None
        for idx, value in enumerate(values, start=1):
            if "штрих" in value and "код" in value:
                current_barcode = idx
            if value == "блюдо":
                current_name = idx
            if "количество" in value and "всего" in value:
                current_qty = idx
        if current_barcode and current_name and current_qty:
            header_row = row_idx
            barcode_col = current_barcode
            name_col = current_name
            qty_col = current_qty
            break

    if not header_row or not barcode_col or not name_col or not qty_col:
        raise ValueError("Не удалось найти обязательные колонки в iiko-файле")

    start, end, days = parse_period(ws["A3"].value)
    aggregated: dict[str, dict[str, Any]] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        barcode_value = ws.cell(row_idx, barcode_col).value
        name_value = ws.cell(row_idx, name_col).value
        qty_value = ws.cell(row_idx, qty_col).value

        if barcode_value in (None, "") and name_value in (None, ""):
            continue
        if barcode_value in (None, ""):
            continue

        barcode = normalize_barcode(barcode_value)
        if not barcode:
            continue

        name = str(name_value or "").strip() or "Без названия"

        try:
            sold_qty = float(qty_value or 0)
        except (ValueError, TypeError):
            continue

        if sold_qty <= 0:
            continue

        entry = aggregated.setdefault(
            barcode,
            {
                "barcode": barcode,
                "name": name,
                "sold_qty": 0,
            },
        )

        if not entry["name"] and name:
            entry["name"] = name

        entry["sold_qty"] += sold_qty

    sales = [
        SaleRow(
            barcode=v["barcode"],
            name=v["name"],
            sold_qty=int(v["sold_qty"]),
        )
        for v in aggregated.values()
    ]

    meta = {
        "start": start,
        "end": end,
        "days": days,
        "total_rows": ws.max_row - header_row,
        "unique_barcodes": len(sales),
    }
    return sales, meta


def write_supplier_order(
    template_path: str | Path,
    output_path: str | Path,
    supplier_name: str,
    buyer_name: str,
    settings: Settings,
    run_date: datetime,
    order_items: list[dict[str, Any]],
) -> None:
    wb = load_workbook(template_path)
    ws = wb.active

    ws["B1"] = run_date.strftime("%Y%m%d%H%M")
    ws["D1"] = run_date.date()
    ws["B2"] = supplier_name
    ws["D2"] = settings.coverage_days
    ws["B3"] = buyer_name

    if settings.delivery_date:
        try:
            ws["B5"] = datetime.fromisoformat(settings.delivery_date).date()
        except ValueError:
            ws["B5"] = settings.delivery_date

    start_row = 8
    max_rows_to_clear = max(len(order_items), 200)

    for row in range(start_row, start_row + max_rows_to_clear):
        ws.cell(row, 1).value = None
        ws.cell(row, 2).value = None
        ws.cell(row, 3).value = None
        ws.cell(row, 4).value = None

    for idx, item in enumerate(order_items, start=1):
        row = start_row + idx - 1
        ws.cell(row, 1).value = idx

        barcode_cell = ws.cell(row, 2)
        barcode_cell.value = item["barcode"]
        barcode_cell.number_format = "@"

        ws.cell(row, 3).value = item["name"]
        ws.cell(row, 4).value = item["orderQty"]

    wb.save(output_path)